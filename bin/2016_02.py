#! /usr/local/bin/python3
""" Check Grades: Spring 2016 version.
"""
import sys
import os
import re
import datetime
from stat import *

# excel stuff
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

# email stuff
import  smtplib
from    email.headerregistry  import Address
from    email.message         import EmailMessage
from    email.utils           import formatdate
from    email.utils           import make_msgid

# CGI stuff -- with debugging
import cgi
import cgitb
import pprint
cgitb.enable()

# oops()
# ---------------------------------------------------------
def oops(msg):
  """ Die because something went oops.
  """
  print("""
   <h1>Unable To Check Grades</h1>
   <h2>{}</h2>
                          """.format(msg))
  exit()

# html2str()
# ---------------------------------------------------------
def html2text(str):
  """ Convert HTML to plain text.
      Change <p> to \n, then get rid of all other tags.
  """
  return_str = str.replace('<p>', '\n')
  return_str = re.sub(r'<.*?>', '', return_str)
  return return_str.replace('\n\n', '\n')

# eval_cell()
# ---------------------------------------------------------
def eval_cell(cell):
  """ Return the value of a cell, even if it contains a
      reference to a cell in another sheet. But formulas
      of other sorts won’t work.
  """
  if cell.data_type == Cell.TYPE_FORMULA:
    sheet, loc = cell.value.split('!')
    sheet = sheet.strip(" ='")
    return wb.get_sheet_by_name(sheet)[loc].value
  return cell.value


# get_student()
# ---------------------------------------------------------
def get_student(wb, sheet_name, student_id):
  """ Get the header row and student data from a named sheet in a
      workbook. Die if the student isn't there.
  """
  try:
    ws = wb.get_sheet_by_name(sheet_name)
  except:
    oops('Workbook sheet {} not found'.format(sheet_name))

  # Get list of active headers
  #   Assume TYPE_NUMERIC is a datetime reference and evaluate it
  headers = []
  c = 1
  while ws.cell(row=1, column=c).value:
    if ws.cell(row=1, column=c).data_type == Cell.TYPE_NUMERIC:
      headers.append(ws.cell(row=1, column=c).value.strftime('%b %e').replace('  ', ' '))
    else:
      headers.append(ws.cell(row=1, column=c).value)
    c += 1
  # find the student
  for r in range(2, ws.max_row+1):
    if ws.cell(row=r, column=1).value == student_id:
      data = [ws.cell(row=r, column=c+1) for c in range(len(headers))]
      # print('get_student: returning ', headers, data)
      return {'headers':headers, 'data':data}
    # else:
    #   print('{}: %{}% does not equal %{}%<br/>'.format(sheet_name, eval_cell(ws.cell(row=r, column=1)), student_id))
  oops('Student ID {} not found in {}'.format(student_id, sheet_name))

# do_sheet()
# ----------------------------------------------------------
def do_sheet(h3_leadin, sheet, text_message, html_message, header_1 = 'Date'):
  """ Generate the email content for one page of the workbook.
      Formatting tends to vary from term to term.
      Col 1: student_id
      Col 2: last_name
      Col 3: first_name
      Col 4-5: If headings are Points and Max, display points/max;
      Remaining columns: display heading and value if heading is not '--'
  """
  headers = sheet['headers']
  data    = sheet['data']
  h3_value  = ''
  if len(headers) > 4 and headers[3] == 'Points' and headers[4] == 'Max':
    h3_value = '{}/{}'.format(data[3].value, data[4].value)
    start_at = 5
  else:
    start_at = 3

  if DEBUG: print(h3_leadin, h3_value, headers, data)

  text = '{}: {}'.format(h3_leadin, h3_value)
  text_message = text_message + text + '\n'
  html_message = html_message + '<h3>{}</h3>'.format(text)
  if len(data) <= start_at:
    return (text_message + 'No information yet\n', html_message + '<p>No information yet</p>')
  text_1 = '{:<7}:'.format(header_1)
  text_2 = '{:<7}:'.format('Score')
  html = '<table><tr><th><strong>{}:<hr/>Score:</strong></th>'.format(header_1)

  num_cols = 0
  for col in range(start_at, len(headers)):
    if headers[col] == '--': continue
    if num_cols > 0:
      if (num_cols % 6) == 0:
        text_1 = text_1 + '\n'
        text_2 = text_2 + '\n'
      if (num_cols % 10) == 0:
        html = html + '</tr><tr><td colspan="11"</td></tr><tr><th><strong>{}:<hr/>Score:</strong></th>'.format(header_1)
    num_cols = num_cols + 1
    header_str = headers[col]
    text_1 = text_1 + '{:8}'.format(header_str)
    if data[col].value:
      text_2 = text_2 + '{:8}'.format(data[col].value)
      html = html + '<td>{}<hr/>{}</td>'.format(header_str.replace(' ','&nbsp;'), data[col].value)
    else:
      text_2 = text_2 + '        '
      html = html + '<td>{}<hr/>{}</td>'.format(header_str.replace(' ','&nbsp;'), '&nbsp;')

  text_message = text_message + text_1 + '\n' + text_2 + '\n'
  html_message = html_message + html + '</tr></table>'
  return (text_message, html_message)


""" Get form data
"""
args = cgi.FieldStorage()

""" Web page headers
"""
sys.stdout.buffer.write("Content-Type: text/html; charset=utf-8\r\n\r\n".encode('utf-8'))

""" Check/Extract form data
"""
if 'course' not in args: oops('Missing course')
course = args['course'].value
sheet_name = course.lower().replace('csci', 'cs').replace('-', '')

if 'semester' not in args: oops('Missing term')
semester = args['semester'].value

if 'student_id' not in args: oops('Missing student ID')

if 'debug' in args:
  DEBUG = True
else:
  DEBUG = False
include_data = ''
if 'include-file' in args:
  with open('../courses/{}/{}/{}'.format(sheet_name, semester, args['include-file'].value),
            encoding='utf-8') as inc:
    include_data = inc.read()

""" Locate and open workbook.
    To provide impenetrable security, we do not reveal the location of the spreadsheet, only its
    name, in error messages.
"""
workbook_file = '{}.xlsx'.format(semester)
workbook_path = '../../Grades/' + workbook_file
if not os.path.isfile(workbook_path): oops('Missing file: ' + workbook_file)

modtime = datetime.datetime.fromtimestamp(os
                                          .stat(workbook_path)
                                          .st_mtime).strftime('%B %d, %Y at %I:%M %p')
try:
  wb = load_workbook(workbook_path, data_only=True)
except:
  oops('Unable to open ' + workbook_file)

# Be sure Student ID supplied is an unambiguous sequence of digits.
student_id = args['student_id'].value.strip()
student_ids = []
try:
    ws = wb.get_sheet_by_name('Roster')
except:
    oops('Workbook sheet "Roster" not found')
for row in range(1, ws.max_row):
  cell = ws.cell(row=row, column=1)
  if cell.data_type == Cell.TYPE_STRING and \
      cell.value.isdigit() and \
      student_id in '{:08}'.format(int(cell.value)):
    student_ids.append('{:08}'.format(int(cell.value)))
if len(student_ids) == 0: oops('Student ID "{}" not in Roster'.format(student_id))
if len(student_ids)  > 1: oops('Student ID "{}" is ambiguous. Use more digits.'.format(student_id))
student_id = student_ids[0]

roster        = get_student(wb, 'Roster', student_id)
takeaways     = get_student(wb, 'Takeaways', student_id)
quizzes       = get_student(wb, 'Quizzes', student_id)
assignments   = get_student(wb, 'Assignments', student_id)
other_grades  = get_student(wb, 'Other Grades', student_id)

data = roster['data']
fname = data[2].value
lname = data[1].value
student_name = '{} {}'.format(fname, lname)
course_score = data[11].value
course_grade = data[12].value
course_message = """
  Your grade for the course is <strong>{}</strong>, which is a <strong>{}</strong>
  """.format(course_score, course_grade)
emails = [data[3].value]
if data[4].value:
  emails.append(data[4].value)
to_list = [Address(student_name, addr_spec=x) for x in emails]

# Construct the HTML and text tables of grades
# --------------------------------------------
text_message = ''
html_message = ''

#  Takeaways, Quizzes, Assignments, and Exams
text_message, html_message = do_sheet('Takeaways',
                                      takeaways,
                                      text_message,
                                      html_message)
text_message, html_message = do_sheet('Quizzes',
                                      quizzes,
                                      text_message,
                                      html_message)
text_message, html_message = do_sheet('Assignments',
                                      assignments,
                                      text_message,
                                      html_message,
                                      header_1 = 'Date/Item')
text_message, html_message = do_sheet('Exams',
                                      other_grades,
                                      text_message,
                                      html_message,
                                      header_1 = 'Item')

""" Style the HTML message
"""
css = """
<style type='text/css'>
  * {font-family: sans-serif;}
  table {
  border-collapse: collapse;
  border: 1px solid lightgray;
  }
  th, td {
    padding: 0.1em 0.5em;
    border: 1px solid lightgray;
    text-align: center;
    vertical-align: top;
  }
  tbody th {
    text-align: right;
  }
  thead {
    background-color: lightgray;
  }
  thead th {
    border: 1px solid black;
  }
  h3 {
    margin:1em 0 0 0;
  }
  .scores {
    display: inline-block;
    width: 5em;
    padding:0.1em;
  }
</style>
"""

""" Localhost page: show the grades table and notes
"""
if 'localhost' in os.environ['SERVER_NAME']:
  email_info = '<h2>Email would go to:</h2><blockquote><p>' + emails[0]
  if len(emails) > 1: email_info = email_info + '<br/>' + emails[1]
  emails_info = email_info + '</blockquote>'

  xhtml_page = """
  <?xml version='1.0' encoding='UTF-8'?>
  <!DOCTYPE html>
  <html xmlns='http://www.w3.org/1999/xhtml' xml:lang='en'>
    <head>
      <title>Check My Grades</title>
      {}
    </head>
    <body>
      <h1>{} Grades for {}</h1>
      <h2>Grades were last updated {}</h2>
      {}
      {}
      {}
    </body>
  </html>
  """.format(css, course, student_name, modtime,
             email_info, html_message, include_data).encode('utf-8')
else:
  email_info  = '<blockquote><p>' + emails[0]
  syntax      = ' with your grades has'
  if len(emails) > 1:
    email_info  = email_info + '<br/>' + emails[1]
    syntax      = 's with your grades have'
  email_info = '<h2>Email{} been sent to:</h2>{}</blockquote>'.format(syntax, email_info)
  xhtml_page = """
  <?xml version='1.0' encoding='UTF-8'?>
  <!DOCTYPE html>
  <html xmlns='http://www.w3.org/1999/xhtml' xml:lang='en'>
    <head>
      <title>Check My Grades</title>
      {}
    </head>
    <body>
      <h1>{} Grades for {}</h1>
      <h2>Grades were last updated {}</h2>
      {}
      {}
    </body>
  </html>
  """.format(css, course, student_name, modtime,
             email_info, include_data).encode('utf-8')
  to_list = [Address(student_name, addr_spec = x) for x in emails]
  text_content  = """
{} Grades for {}
Grades were last updated {}
{}
{}
""".format(course, student_name, modtime, text_message, html2text(include_data))

  html_content  = """
  <head>
    {}
  </head>
  <body>
    <h1>{} Grades for {}</h1>
    <h2>Grades were last updated {}</h2>
    {}
    {}
  </body>
</html>
""".format(css, course, student_name, modtime, html_message, include_data)

  msg               = EmailMessage()
  msg['Subject']    = 'Your CSCI-100 Grades'
  msg['From']       = Address('Christopher Vickery', addr_spec='christopher.vickery@qc.cuny.edu')
  msg['To']         = to_list
  msg['Bcc']        = Address('Christopher Vickery', addr_spec='christopher.vickery@qc.cuny.edu')
  msg.add_header('Reply-To',    'christopher.vickery@qc.cuny.edu')
  msg.add_header('Date',        formatdate(localtime=True))
  msg.add_header('Message-ID',  make_msgid())
  msg.set_content(text_message)
  msg.add_alternative(html_content, subtype='html')
  mailer = smtplib.SMTP('smtp.qc.cuny.edu')
  mailer.send_message(msg)
  mailer.quit()
sys.stdout.buffer.write(xhtml_page)
